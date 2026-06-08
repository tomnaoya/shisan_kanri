import os
import io
import csv
from datetime import timedelta, datetime

from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required, get_jwt_identity
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# App Configuration
# ---------------------------------------------------------------------------
app = Flask(__name__)

DATA_DIR = os.environ.get("DATA_DIR", "/var/data")
try:
    os.makedirs(DATA_DIR, exist_ok=True)
except PermissionError:
    DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(DATA_DIR, exist_ok=True)

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DATA_DIR}/assets.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["JWT_SECRET_KEY"] = os.environ.get("JWT_SECRET_KEY", "change-me-in-production")
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=8)

db = SQLAlchemy(app)
jwt = JWTManager(app)

CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=False)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Management code: <CategoryLetter><LocationDigit>-NNNNN
CATEGORY_LETTER = {
    "medical":        "A",
    "medical_supply": "B",
    "electronic":     "C",
    "software":       "D",
    "equipment":      "E",
    "intangible":     "N",
    "other":          "Z",
}

LOCATION_DIGIT = {
    "豊洲院":              "1",
    "勝どき院":            "2",
    "田町芝浦院":          "3",
    "ガーデン院小児耳鼻":  "4",
    "ガーデン院皮膚":      "5",
    "柏院":                "6",
    "サポートチーム":      "0",
}

CATEGORY_LABELS = {
    "medical":        "医療機器",
    "medical_supply": "医療資材",
    "electronic":     "電子機器",
    "software":       "ソフトウェア",
    "equipment":      "器具備品",
    "intangible":     "非実在資産",
    "other":          "その他",
}

CATEGORY_REVERSE = {v: k for k, v in CATEGORY_LABELS.items()}

# Maintenance options: "有", "無", "不明"
MAINTENANCE_VALUES = {"有", "無", "不明"}

# 閲覧を所属院に制限するスイッチ。
# True  : 一般ユーザーは自分の所属院の資産のみ閲覧 (admin / サポートチームは全件)
# False : 全ユーザーが全資産を閲覧できる旧仕様に戻る
# ↓ 旧仕様に戻したい場合はこの行を False にする (またはコメントアウトして次行を有効化)
RESTRICT_VIEW_BY_LOCATION = True
# RESTRICT_VIEW_BY_LOCATION = False  # ← 旧仕様(全件閲覧)に戻すときはこちらを有効化

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    display_name = db.Column(db.String(120), nullable=True)
    role = db.Column(db.String(20), nullable=False, default="user")  # admin / user
    location = db.Column(db.String(100), nullable=True)  # belonging location (null = all for admin)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {
            "id": self.id,
            "username": self.username,
            "display_name": self.display_name,
            "role": self.role,
            "location": self.location,
            "created_at": self.created_at.isoformat() if self.created_at else None,
        }


class Asset(db.Model):
    __tablename__ = "assets"
    id = db.Column(db.Integer, primary_key=True)
    management_code = db.Column(db.String(100), unique=True, nullable=False, index=True)
    name = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(50), nullable=False)
    category_other = db.Column(db.String(200), nullable=True)
    location = db.Column(db.String(100), nullable=False)
    location_other = db.Column(db.String(200), nullable=True)
    department = db.Column(db.String(100), nullable=False)
    department_other = db.Column(db.String(200), nullable=True)
    purchase_from = db.Column(db.String(200), nullable=True)
    purchase_price = db.Column(db.Integer, nullable=True)
    purchase_date = db.Column(db.String(20), nullable=True)
    maintenance_status = db.Column(db.String(10), default="無")  # 有 / 無 / 不明
    operating_status = db.Column(db.String(10), default="稼働中")  # 稼働中 / 休眠 / 廃棄済
    disposed_date = db.Column(db.String(20), nullable=True)  # 廃棄年月日
    maintenance_info = db.Column(db.Text, nullable=True)
    maintenance_link = db.Column(db.String(500), nullable=True)
    depreciation_period_months = db.Column(db.Integer, nullable=True)
    lease_period_months = db.Column(db.Integer, nullable=True)
    manager = db.Column(db.String(100), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    is_deleted = db.Column(db.Boolean, default=False, index=True)
    deleted_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Migration helper: old has_maintenance boolean
    has_maintenance = db.Column(db.Boolean, nullable=True)

    def to_dict(self):
        return {
            "id": self.id,
            "management_code": self.management_code,
            "name": self.name,
            "category": self.category,
            "category_other": self.category_other,
            "location": self.location,
            "location_other": self.location_other,
            "department": self.department,
            "department_other": self.department_other,
            "purchase_from": self.purchase_from,
            "purchase_price": self.purchase_price,
            "purchase_date": self.purchase_date,
            "maintenance_status": self.maintenance_status or "無",
            "operating_status": self.operating_status or "稼働中",
            "disposed_date": self.disposed_date,
            "maintenance_info": self.maintenance_info,
            "maintenance_link": self.maintenance_link,
            "depreciation_period_months": self.depreciation_period_months,
            "lease_period_months": self.lease_period_months,
            "manager": self.manager,
            "notes": self.notes,
            "is_deleted": self.is_deleted,
            "created_at": self.created_at.isoformat() if self.created_at else None,
            "updated_at": self.updated_at.isoformat() if self.updated_at else None,
        }


class Department(db.Model):
    __tablename__ = "departments"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {"id": self.id, "name": self.name, "is_default": self.is_default}


class Location(db.Model):
    __tablename__ = "locations"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {"id": self.id, "name": self.name, "is_default": self.is_default}


# ---------------------------------------------------------------------------
# Auto-numbering
# ---------------------------------------------------------------------------

def generate_management_code(category, location):
    """Generate next management_code for (category, location).

    Format: <CategoryLetter><LocationDigit>-NNNNN
    Unknown category -> Z, unknown location -> 0.
    """
    letter = CATEGORY_LETTER.get(category, "Z")
    digit = LOCATION_DIGIT.get(location, "0")
    prefix = f"{letter}{digit}"
    like_pattern = f"{prefix}-%"
    last = (
        Asset.query
        .filter(Asset.management_code.like(like_pattern))
        .order_by(Asset.management_code.desc())
        .first()
    )
    if last:
        try:
            num = int(last.management_code.split("-")[1]) + 1
        except (IndexError, ValueError):
            num = 1
    else:
        num = 1
    return f"{prefix}-{num:05d}"


# ---------------------------------------------------------------------------
# Seed Data
# ---------------------------------------------------------------------------

DEFAULT_DEPARTMENTS = [
    "小児科1診", "小児科2診", "耳鼻科1診", "耳鼻科2診",
    "皮膚科診察室", "バックヤード", "受付", "その他",
]

DEFAULT_LOCATIONS = [
    "豊洲院", "勝どき院", "田町芝浦院", "ガーデン院小児耳鼻",
    "ガーデン院皮膚", "柏院", "サポートチーム",
]


def seed_data():
    if User.query.count() == 0:
        admin = User(username="admin", display_name="管理者", role="admin", location=None)
        admin.set_password("admin")
        db.session.add(admin)

    if Department.query.count() == 0:
        for name in DEFAULT_DEPARTMENTS:
            db.session.add(Department(name=name, is_default=True))

    if Location.query.count() == 0:
        for name in DEFAULT_LOCATIONS:
            db.session.add(Location(name=name, is_default=True))

    db.session.commit()


def migrate_data():
    """Migrate existing data to new schema."""
    # 1. Add role to existing users
    for user in User.query.all():
        if not user.role:
            user.role = "admin"

    # 2. Migrate has_maintenance -> maintenance_status
    for asset in Asset.query.all():
        if asset.maintenance_status is None or asset.maintenance_status == "":
            if asset.has_maintenance is True:
                asset.maintenance_status = "有"
            elif asset.has_maintenance is False:
                asset.maintenance_status = "無"
            else:
                asset.maintenance_status = "不明"
        if asset.is_deleted is None:
            asset.is_deleted = False

    db.session.flush()

    # 3. Location merge: 有明院 -> ガーデン院小児耳鼻, 有明ひふか院 -> ガーデン院皮膚
    LOCATION_MERGE = {
        "有明院": "ガーデン院小児耳鼻",
        "有明ひふか院": "ガーデン院皮膚",
    }
    for old_loc, new_loc in LOCATION_MERGE.items():
        Asset.query.filter_by(location=old_loc).update({"location": new_loc})
    # Remove merged locations from Location table
    for old_loc in LOCATION_MERGE:
        Location.query.filter_by(name=old_loc).delete()

    db.session.commit()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_current_user():
    uid = int(get_jwt_identity())
    return User.query.get(uid)


def check_location_permission(user, asset_location):
    """Check if user has permission for given location."""
    if user.role == "admin":
        return True
    if user.location == "サポートチーム":
        return True
    return user.location == asset_location


def restrict_assets_to_user(query, user):
    """Limit an Asset query to assets the user is allowed to view.

    admin / サポートチーム see all assets; a regular user sees only
    assets in their own belonging location.
    RESTRICT_VIEW_BY_LOCATION を False にすると無効化され全件返す。
    """
    if not RESTRICT_VIEW_BY_LOCATION:
        return query
    if user.role == "admin" or user.location == "サポートチーム":
        return query
    return query.filter(Asset.location == user.location)


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    user = User.query.filter_by(username=username).first()
    if not user or not user.check_password(password):
        return jsonify({"error": "ユーザー名またはパスワードが正しくありません"}), 401

    token = create_access_token(identity=str(user.id))
    return jsonify({"access_token": token, "user": user.to_dict()})


@app.route("/api/auth/me", methods=["GET"])
@jwt_required()
def me():
    user = get_current_user()
    if not user:
        return jsonify({"error": "ユーザーが見つかりません"}), 404
    return jsonify(user.to_dict())


# ---------------------------------------------------------------------------
# Asset CRUD
# ---------------------------------------------------------------------------

@app.route("/api/assets", methods=["GET"])
@jwt_required()
def list_assets():
    user = get_current_user()
    q = Asset.query.filter_by(is_deleted=False)

    # 所属院でフィルタ (admin / サポートチームは全件)
    q = restrict_assets_to_user(q, user)

    # Filters
    category = request.args.get("category")
    location = request.args.get("location")
    search = request.args.get("search")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    if category:
        q = q.filter(Asset.category == category)
    if location:
        q = q.filter(Asset.location == location)
    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(
                Asset.name.ilike(like),
                Asset.management_code.ilike(like),
                Asset.manager.ilike(like),
                Asset.purchase_from.ilike(like),
            )
        )
    if date_from:
        q = q.filter(Asset.purchase_date >= date_from)
    if date_to:
        q = q.filter(Asset.purchase_date <= date_to)
    op_status = request.args.getlist("operating_status")
    if op_status:
        q = q.filter(Asset.operating_status.in_(op_status))

    # Sort
    sort = request.args.get("sort", "updated_at")
    order = request.args.get("order", "desc")
    sort_col = getattr(Asset, sort, Asset.updated_at)
    q = q.order_by(sort_col.desc() if order == "desc" else sort_col.asc())

    # Paginate
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "items": [a.to_dict() for a in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "pages": pagination.pages,
    })


@app.route("/api/assets/next-code/<category>", methods=["GET"])
@jwt_required()
def next_code(category):
    """Preview next management code for a (category, location)."""
    location = request.args.get("location", "")
    code = generate_management_code(category, location)
    return jsonify({"code": code})


@app.route("/api/assets", methods=["POST"])
@jwt_required()
def create_asset():
    user = get_current_user()
    data = request.get_json(silent=True) or {}

    required = ["name", "category", "location"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"必須項目が不足しています: {', '.join(missing)}"}), 400

    loc = data["location"]
    if not check_location_permission(user, loc):
        return jsonify({"error": "この設置場所への登録権限がありません"}), 403

    # Validate
    if loc == "その他" and not data.get("location_other"):
        return jsonify({"error": "設置場所「その他」の場合は詳細を入力してください"}), 400
    if data.get("category") == "other" and not data.get("category_other"):
        return jsonify({"error": "種別「その他」の場合は詳細を入力してください"}), 400

    dept = data.get("department", "その他")
    if not dept:
        dept = "その他"

    # Auto-generate management code (category + location)
    code = generate_management_code(data["category"], loc)

    maint = data.get("maintenance_status", "無")
    if maint not in MAINTENANCE_VALUES:
        maint = "無"

    op_st = data.get("operating_status", "稼働中")
    if op_st == "廃棄済" and not data.get("disposed_date"):
        return jsonify({"error": "稼働状況が「廃棄済」の場合は廃棄年月日を入力してください"}), 400

    asset = Asset(
        management_code=code,
        name=data["name"],
        category=data["category"],
        category_other=data.get("category_other"),
        location=loc,
        location_other=data.get("location_other"),
        department=dept,
        department_other=data.get("department_other"),
        purchase_from=data.get("purchase_from"),
        purchase_price=data.get("purchase_price"),
        purchase_date=data.get("purchase_date"),
        maintenance_status=maint,
        operating_status=op_st,
        disposed_date=data.get("disposed_date"),
        maintenance_info=data.get("maintenance_info"),
        maintenance_link=data.get("maintenance_link"),
        depreciation_period_months=data.get("depreciation_period_months"),
        lease_period_months=data.get("lease_period_months"),
        manager=data.get("manager"),
        notes=data.get("notes"),
        is_deleted=False,
    )
    db.session.add(asset)
    db.session.commit()
    return jsonify(asset.to_dict()), 201


@app.route("/api/assets/<int:asset_id>", methods=["GET"])
@jwt_required()
def get_asset(asset_id):
    user = get_current_user()
    asset = Asset.query.get_or_404(asset_id)
    if RESTRICT_VIEW_BY_LOCATION and not check_location_permission(user, asset.location):
        return jsonify({"error": "この資産の閲覧権限がありません"}), 403
    return jsonify(asset.to_dict())


@app.route("/api/assets/<int:asset_id>", methods=["PUT"])
@jwt_required()
def update_asset(asset_id):
    user = get_current_user()
    asset = Asset.query.get_or_404(asset_id)

    if asset.is_deleted:
        return jsonify({"error": "削除済みの資産は編集できません"}), 400

    if not check_location_permission(user, asset.location):
        return jsonify({"error": "この資産の編集権限がありません"}), 403

    data = request.get_json(silent=True) or {}

    # 管理番号は登録後変更しない (種別変更・院移動でも維持)

    # Validate location change permission
    new_loc = data.get("location", asset.location)
    if new_loc != asset.location and not check_location_permission(user, new_loc):
        return jsonify({"error": "移動先の設置場所への権限がありません"}), 403

    # Validate disposed_date
    new_op = data.get("operating_status", asset.operating_status)
    if new_op == "廃棄済" and not (data.get("disposed_date") or asset.disposed_date):
        return jsonify({"error": "稼働状況が「廃棄済」の場合は廃棄年月日を入力してください"}), 400

    updatable = [
        "name", "category", "category_other",
        "location", "location_other", "department", "department_other",
        "purchase_from", "purchase_price", "purchase_date",
        "maintenance_status", "maintenance_info", "maintenance_link",
        "operating_status", "disposed_date",
        "depreciation_period_months", "lease_period_months",
        "manager", "notes",
    ]
    for field in updatable:
        if field in data:
            setattr(asset, field, data[field])

    db.session.commit()
    return jsonify(asset.to_dict())


@app.route("/api/assets/<int:asset_id>", methods=["DELETE"])
@jwt_required()
def delete_asset(asset_id):
    """Soft delete – mark as deleted, never physically remove."""
    user = get_current_user()
    asset = Asset.query.get_or_404(asset_id)

    if not check_location_permission(user, asset.location):
        return jsonify({"error": "この資産の削除権限がありません"}), 403

    asset.is_deleted = True
    asset.deleted_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"message": "削除しました"}), 200


# ---------------------------------------------------------------------------
# Bulk Download
# ---------------------------------------------------------------------------

EXCEL_HEADERS = [
    ("管理番号", "management_code"),
    ("資産名", "name"),
    ("種別", "_category_display"),
    ("設置場所", "_location_display"),
    ("設置部室", "_department_display"),
    ("購入元", "purchase_from"),
    ("購入金額", "purchase_price"),
    ("購入日", "purchase_date"),
    ("保守", "_maintenance_display"),
    ("稼働状況", "operating_status"),
    ("廃棄年月日", "disposed_date"),
    ("保守情報", "maintenance_info"),
    ("保守リンク", "maintenance_link"),
    ("減価償却期間(月)", "depreciation_period_months"),
    ("リース期間(月)", "lease_period_months"),
    ("管理者", "manager"),
    ("備考", "notes"),
    ("登録日", "created_at"),
    ("更新日", "updated_at"),
]


def _asset_row(asset):
    d = asset.to_dict()
    d["_category_display"] = CATEGORY_LABELS.get(d["category"], d.get("category_other") or d["category"])
    loc = d["location"]
    if loc == "その他" and d.get("location_other"):
        loc = f"その他({d['location_other']})"
    d["_location_display"] = loc
    dept = d["department"]
    if dept == "その他" and d.get("department_other"):
        dept = f"その他({d['department_other']})"
    d["_department_display"] = dept
    d["_maintenance_display"] = d.get("maintenance_status", "無")
    return d


@app.route("/api/assets/download", methods=["GET"])
@jwt_required()
def download_assets():
    user = get_current_user()
    fmt = request.args.get("format", "xlsx")
    q = Asset.query.filter_by(is_deleted=False)
    q = restrict_assets_to_user(q, user)
    assets = q.order_by(Asset.management_code).all()

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([h[0] for h in EXCEL_HEADERS])
        for asset in assets:
            row = _asset_row(asset)
            writer.writerow([row.get(h[1], "") for h in EXCEL_HEADERS])
        output.seek(0)
        buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        return send_file(buf, mimetype="text/csv", as_attachment=True,
                         download_name=f"assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "資産一覧"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, (header, _) in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_idx, asset in enumerate(assets, 2):
        row_data = _asset_row(asset)
        for col_idx, (_, key) in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
            cell.border = thin_border
    for col_idx, (header, _) in enumerate(EXCEL_HEADERS, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(len(header) * 2.5, 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


# ---------------------------------------------------------------------------
# CSV Import
# ---------------------------------------------------------------------------

IMPORT_HEADERS = [
    "管理番号", "資産名", "種別", "種別(その他)", "設置場所", "設置場所(その他)",
    "設置部室", "設置部室(その他)", "購入元", "購入金額", "購入日",
    "保守有無", "保守情報", "保守リンク", "減価償却期間(月)", "リース期間(月)",
    "管理者", "備考",
]


@app.route("/api/assets/upload-template", methods=["GET"])
@jwt_required()
def download_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(IMPORT_HEADERS)
    writer.writerow([
        "(自動採番)", "超音波診断装置", "医療機器", "", "豊洲院", "",
        "小児科1診", "", "GEヘルスケア", "4800000", "2024-04-15",
        "有", "年間保守契約", "https://example.com", "72", "",
        "佐藤医師", "備考メモ",
    ])
    output.seek(0)
    buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name="import_template.csv")


@app.route("/api/assets/upload", methods=["POST"])
@jwt_required()
def upload_assets():
    user = get_current_user()

    if "file" not in request.files:
        return jsonify({"error": "ファイルが選択されていません"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".csv"):
        return jsonify({"error": "CSVファイルのみ対応しています"}), 400

    try:
        raw = file.read()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("shift_jis")

        reader = csv.DictReader(io.StringIO(text))
        created = 0
        skipped = 0
        errors = []

        for i, row in enumerate(reader, start=2):
            name = (row.get("資産名") or "").strip()
            if not name:
                skipped += 1
                continue

            cat_label = (row.get("種別") or "").strip()
            category = CATEGORY_REVERSE.get(cat_label, "other")
            category_other = (row.get("種別(その他)") or "").strip()
            if cat_label and cat_label not in CATEGORY_REVERSE:
                category = "other"
                category_other = cat_label

            location = (row.get("設置場所") or "").strip()
            location_other = (row.get("設置場所(その他)") or "").strip()
            if not location:
                errors.append(f"行{i}: 設置場所が空です")
                skipped += 1
                continue

            if not check_location_permission(user, location):
                errors.append(f"行{i}: {location}への登録権限がありません")
                skipped += 1
                continue

            department = (row.get("設置部室") or "").strip()
            department_other = (row.get("設置部室(その他)") or "").strip()
            if not department:
                department = "その他"

            def parse_int(val):
                val = (val or "").strip().replace(",", "")
                if not val:
                    return None
                try:
                    return int(float(val))
                except (ValueError, TypeError):
                    return None

            maint_str = (row.get("保守有無") or "").strip()
            if maint_str in ("有", "あり", "○", "1", "true", "True", "YES", "yes"):
                maint = "有"
            elif maint_str in ("不明",):
                maint = "不明"
            else:
                maint = "無"

            # Auto-generate code (category + location)
            code = generate_management_code(category, location)

            # Put old management_code (from CSV) into notes
            old_code = (row.get("管理番号") or "").strip()
            notes_val = (row.get("備考") or "").strip()
            if old_code:
                if notes_val:
                    notes_val = f"旧管理番号: {old_code} / {notes_val}"
                else:
                    notes_val = f"旧管理番号: {old_code}"

            asset = Asset(
                management_code=code,
                name=name,
                category=category,
                category_other=category_other or None,
                location=location,
                location_other=location_other or None,
                department=department,
                department_other=department_other or None,
                purchase_from=(row.get("購入元") or "").strip() or None,
                purchase_price=parse_int(row.get("購入金額")),
                purchase_date=(row.get("購入日") or "").strip() or None,
                maintenance_status=maint,
                operating_status="稼働中",
                maintenance_info=(row.get("保守情報") or "").strip() or None,
                maintenance_link=(row.get("保守リンク") or "").strip() or None,
                depreciation_period_months=parse_int(row.get("減価償却期間(月)")),
                lease_period_months=parse_int(row.get("リース期間(月)")),
                manager=(row.get("管理者") or "").strip() or None,
                notes=notes_val or None,
                is_deleted=False,
            )
            db.session.add(asset)
            created += 1

        db.session.commit()
        return jsonify({
            "message": f"{created}件を登録しました",
            "created": created,
            "skipped": skipped,
            "errors": errors[:20],
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"CSVの処理中にエラーが発生しました: {str(e)}"}), 400


# ---------------------------------------------------------------------------
# Department CRUD
# ---------------------------------------------------------------------------

@app.route("/api/departments", methods=["GET"])
@jwt_required()
def list_departments():
    depts = Department.query.order_by(Department.id).all()
    return jsonify([d.to_dict() for d in depts])


@app.route("/api/departments", methods=["POST"])
@jwt_required()
def create_department():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "部室名を入力してください"}), 400
    if Department.query.filter_by(name=name).first():
        return jsonify({"error": "同名の部室が既に存在します"}), 409
    dept = Department(name=name, is_default=False)
    db.session.add(dept)
    db.session.commit()
    return jsonify(dept.to_dict()), 201


@app.route("/api/departments/<int:dept_id>", methods=["DELETE"])
@jwt_required()
def delete_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    db.session.delete(dept)
    db.session.commit()
    return jsonify({"message": "削除しました"}), 200


# ---------------------------------------------------------------------------
# Location CRUD (admin only)
# ---------------------------------------------------------------------------

@app.route("/api/locations", methods=["GET"])
@jwt_required()
def list_locations():
    locs = Location.query.order_by(Location.id).all()
    return jsonify([l.to_dict() for l in locs])


@app.route("/api/locations", methods=["POST"])
@jwt_required()
def create_location():
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "設置場所名を入力してください"}), 400
    if Location.query.filter_by(name=name).first():
        return jsonify({"error": "同名の設置場所が既に存在します"}), 409
    loc = Location(name=name, is_default=False)
    db.session.add(loc)
    db.session.commit()
    return jsonify(loc.to_dict()), 201


@app.route("/api/locations/<int:loc_id>", methods=["PUT"])
@jwt_required()
def update_location(loc_id):
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    loc = Location.query.get_or_404(loc_id)
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if name and name != loc.name:
        if Location.query.filter_by(name=name).first():
            return jsonify({"error": "同名の設置場所が既に存在します"}), 409
        loc.name = name
    db.session.commit()
    return jsonify(loc.to_dict())


@app.route("/api/locations/<int:loc_id>", methods=["DELETE"])
@jwt_required()
def delete_location(loc_id):
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    loc = Location.query.get_or_404(loc_id)
    db.session.delete(loc)
    db.session.commit()
    return jsonify({"message": "削除しました"}), 200


# ---------------------------------------------------------------------------
# User Management (admin only)
# ---------------------------------------------------------------------------

@app.route("/api/users", methods=["GET"])
@jwt_required()
def list_users():
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    users = User.query.order_by(User.id).all()
    return jsonify([u.to_dict() for u in users])


@app.route("/api/users", methods=["POST"])
@jwt_required()
def create_user():
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return jsonify({"error": "ユーザー名とパスワードは必須です"}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"error": "このユーザー名は既に使用されています"}), 409
    new_user = User(
        username=username,
        display_name=data.get("display_name", username),
        role=data.get("role", "user"),
        location=data.get("location"),
    )
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    return jsonify(new_user.to_dict()), 201


@app.route("/api/users/<int:uid>", methods=["PUT"])
@jwt_required()
def update_user(uid):
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    target = User.query.get_or_404(uid)
    data = request.get_json(silent=True) or {}

    if "display_name" in data:
        target.display_name = data["display_name"]
    if "role" in data:
        target.role = data["role"]
    if "location" in data:
        target.location = data["location"]
    if data.get("password"):
        target.set_password(data["password"])

    db.session.commit()
    return jsonify(target.to_dict())


@app.route("/api/users/<int:uid>", methods=["DELETE"])
@jwt_required()
def delete_user(uid):
    user = get_current_user()
    if user.role != "admin":
        return jsonify({"error": "管理者権限が必要です"}), 403
    if uid == user.id:
        return jsonify({"error": "自分自身は削除できません"}), 400
    target = User.query.get_or_404(uid)
    db.session.delete(target)
    db.session.commit()
    return jsonify({"message": "削除しました"}), 200


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

@app.route("/api/stats", methods=["GET"])
@jwt_required()
def stats():
    user = get_current_user()
    q = Asset.query.filter_by(is_deleted=False)

    # 所属院でフィルタ (admin / サポートチームは全件)
    q = restrict_assets_to_user(q, user)

    total = q.count()
    by_category = q.with_entities(Asset.category, db.func.count(Asset.id)).group_by(Asset.category).all()
    by_location = q.with_entities(Asset.location, db.func.count(Asset.id)).group_by(Asset.location).all()
    maintenance_count = q.filter(Asset.maintenance_status == "有").count()

    return jsonify({
        "total": total,
        "by_category": {CATEGORY_LABELS.get(c, c): n for c, n in by_category},
        "by_location": dict(by_location),
        "maintenance_count": maintenance_count,
    })


# ---------------------------------------------------------------------------
# Health & Frontend
# ---------------------------------------------------------------------------

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "timestamp": datetime.utcnow().isoformat()})


@app.route("/")
def serve_frontend():
    return send_file("index.html")


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------

with app.app_context():
    db.create_all()

    # Migrate existing tables: add missing columns
    import sqlite3 as _sqlite3
    _conn = _sqlite3.connect(f"{DATA_DIR}/assets.db")
    _cur = _conn.cursor()

    def _add_col(table, column, col_type, default=None):
        try:
            _cur.execute(f"SELECT {column} FROM {table} LIMIT 1")
        except _sqlite3.OperationalError:
            dc = f" DEFAULT '{default}'" if default is not None else ""
            _cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}{dc}")

    _add_col("users", "role", "VARCHAR(20)", "admin")
    _add_col("users", "location", "VARCHAR(100)")
    _add_col("assets", "maintenance_status", "VARCHAR(10)", "無")
    _add_col("assets", "is_deleted", "BOOLEAN", "0")
    _add_col("assets", "deleted_at", "DATETIME")
    _add_col("assets", "has_maintenance", "BOOLEAN")
    _add_col("assets", "operating_status", "VARCHAR(10)", "稼働中")
    _add_col("assets", "disposed_date", "VARCHAR(20)")
    _conn.commit()

    # ---- One-shot 再採番 v2: <CategoryLetter><LocationDigit>-NNNNN ----
    # 多重実行防止: DATA_DIR の .renumber_v2_done フラグで一回限り
    # 本番稼働後はこのブロックは絶対に走らない (フラグ削除しない限り)
    _renumber_flag = os.path.join(DATA_DIR, ".renumber_v2_done")
    if not os.path.exists(_renumber_flag):
        # 一旦すべて TMP_ 接頭辞でリネームして UNIQUE 制約衝突を回避
        _cur.execute("UPDATE assets SET management_code = 'TMP_' || management_code WHERE management_code NOT LIKE 'TMP_%'")

        # (category, location) ごとに id 順で 1 から採番
        for _cat, _letter in [("medical","A"),("medical_supply","B"),("electronic","C"),
                              ("software","D"),("equipment","E"),("intangible","N"),("other","Z")]:
            for _loc, _digit in [("豊洲院","1"),("勝どき院","2"),("田町芝浦院","3"),
                                 ("ガーデン院小児耳鼻","4"),("ガーデン院皮膚","5"),
                                 ("柏院","6"),("サポートチーム","0")]:
                _prefix = f"{_letter}{_digit}"
                _cur.execute(
                    "SELECT id, management_code, notes FROM assets "
                    "WHERE category=? AND location=? ORDER BY id",
                    (_cat, _loc),
                )
                _rows = _cur.fetchall()
                for _idx, (_id, _tmp_code, _old_notes) in enumerate(_rows, start=1):
                    _new_code = f"{_prefix}-{_idx:05d}"
                    _real_old = (_tmp_code or "").replace("TMP_", "", 1)
                    _note = _old_notes or ""
                    if _real_old and _real_old != _new_code and _real_old not in _note:
                        _note = f"旧管理番号: {_real_old} / {_note}" if _note else f"旧管理番号: {_real_old}"
                    _cur.execute("UPDATE assets SET management_code=?, notes=? WHERE id=?",
                                 (_new_code, _note, _id))

        # 未分類 (location/category がマスタに無い) は Z0- で連番
        _cur.execute("SELECT management_code FROM assets WHERE management_code LIKE 'Z0-%' "
                     "ORDER BY management_code DESC LIMIT 1")
        _max = _cur.fetchone()
        try:
            _next_z0 = int(_max[0].split("-")[1]) + 1 if _max else 1
        except (IndexError, ValueError):
            _next_z0 = 1
        _cur.execute("SELECT id, management_code, notes FROM assets WHERE management_code LIKE 'TMP_%'")
        for _id, _tmp_code, _old_notes in _cur.fetchall():
            _new_code = f"Z0-{_next_z0:05d}"
            _next_z0 += 1
            _real_old = (_tmp_code or "").replace("TMP_", "", 1)
            _note = _old_notes or ""
            if _real_old and _real_old != _new_code and _real_old not in _note:
                _note = f"旧管理番号: {_real_old} / {_note}" if _note else f"旧管理番号: {_real_old}"
            _cur.execute("UPDATE assets SET management_code=?, notes=? WHERE id=?",
                         (_new_code, _note, _id))

        _conn.commit()
        with open(_renumber_flag, "w") as _f:
            _f.write(datetime.utcnow().isoformat())

    # ---- One-shot 再採番 v3: 種別変更を反映して全件再採番 ----
    # v2 と同じロジック。種別をユーザーが更新したため再度走らせる。
    _renumber_v3_flag = os.path.join(DATA_DIR, ".renumber_v3_done")
    if not os.path.exists(_renumber_v3_flag):
        _cur.execute("UPDATE assets SET management_code = 'TMP_' || management_code WHERE management_code NOT LIKE 'TMP_%'")

        for _cat, _letter in [("medical","A"),("medical_supply","B"),("electronic","C"),
                              ("software","D"),("equipment","E"),("intangible","N"),("other","Z")]:
            for _loc, _digit in [("豊洲院","1"),("勝どき院","2"),("田町芝浦院","3"),
                                 ("ガーデン院小児耳鼻","4"),("ガーデン院皮膚","5"),
                                 ("柏院","6"),("サポートチーム","0")]:
                _prefix = f"{_letter}{_digit}"
                _cur.execute(
                    "SELECT id, management_code, notes FROM assets "
                    "WHERE category=? AND location=? ORDER BY id",
                    (_cat, _loc),
                )
                _rows = _cur.fetchall()
                for _idx, (_id, _tmp_code, _old_notes) in enumerate(_rows, start=1):
                    _new_code = f"{_prefix}-{_idx:05d}"
                    _real_old = (_tmp_code or "").replace("TMP_", "", 1)
                    _note = _old_notes or ""
                    if _real_old and _real_old != _new_code and _real_old not in _note:
                        _note = f"旧管理番号: {_real_old} / {_note}" if _note else f"旧管理番号: {_real_old}"
                    _cur.execute("UPDATE assets SET management_code=?, notes=? WHERE id=?",
                                 (_new_code, _note, _id))

        _cur.execute("SELECT management_code FROM assets WHERE management_code LIKE 'Z0-%' "
                     "ORDER BY management_code DESC LIMIT 1")
        _max = _cur.fetchone()
        try:
            _next_z0 = int(_max[0].split("-")[1]) + 1 if _max else 1
        except (IndexError, ValueError):
            _next_z0 = 1
        _cur.execute("SELECT id, management_code, notes FROM assets WHERE management_code LIKE 'TMP_%'")
        for _id, _tmp_code, _old_notes in _cur.fetchall():
            _new_code = f"Z0-{_next_z0:05d}"
            _next_z0 += 1
            _real_old = (_tmp_code or "").replace("TMP_", "", 1)
            _note = _old_notes or ""
            if _real_old and _real_old != _new_code and _real_old not in _note:
                _note = f"旧管理番号: {_real_old} / {_note}" if _note else f"旧管理番号: {_real_old}"
            _cur.execute("UPDATE assets SET management_code=?, notes=? WHERE id=?",
                         (_new_code, _note, _id))

        _conn.commit()
        with open(_renumber_v3_flag, "w") as _f:
            _f.write(datetime.utcnow().isoformat())

    # ---- One-shot: 備考欄から「現行採番体系の旧管理番号」のみ削除 ----
    # v2/v3 で追記された「旧管理番号: XX-NNNNN」(現行体系) を取り除き、
    # 元々の管理番号 (例: MED-0342) は備考に残す。
    _cleanup_flag = os.path.join(DATA_DIR, ".cleanup_renumbered_old_codes_done")
    if not os.path.exists(_cleanup_flag):
        import re as _re
        _current_code_re = _re.compile(r"^[ABCDENZ][0-6]-\d{5}$")
        _cur.execute("SELECT id, notes FROM assets WHERE notes LIKE '%旧管理番号:%'")
        for _id, _notes in _cur.fetchall():
            if not _notes:
                continue
            _segments = [s.strip() for s in _notes.split(" / ")]
            _kept = []
            for _seg in _segments:
                if _seg.startswith("旧管理番号:"):
                    _code = _seg[len("旧管理番号:"):].strip()
                    if _current_code_re.match(_code):
                        continue
                _kept.append(_seg)
            _new_notes = " / ".join(_kept) if _kept else None
            if _new_notes != _notes:
                _cur.execute("UPDATE assets SET notes=? WHERE id=?", (_new_notes, _id))
        _conn.commit()
        with open(_cleanup_flag, "w") as _f:
            _f.write(datetime.utcnow().isoformat())

    _conn.close()

    seed_data()
    migrate_data()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
